#!/usr/bin/env python3

import requests
import os
import re
import json
import logging
from datetime import datetime
from collections import Counter
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Selector to choose the function to run
# Options: "fetch", "process"
mode = "process"  # Default mode is set to "process"

# Base directory for all file paths
base_directory = "/Users/phobrla/Library/CloudStorage/OneDrive-Personal/Documents/Bulgarian Language Learning"

# Paths and directories
input_file_path = os.path.join(base_directory, "Inputs_for_PONS_API.txt")
output_directory = os.path.join(base_directory, "PONS json Files")
concatenated_file_path = os.path.join(output_directory, "concatenated.json")
query_parts_of_speech_json_path = os.path.join(base_directory, "Query Parts of Speech.json")
flashcards_xlsm_path = os.path.join(base_directory, "Flashcards.xlsm")

# Ensure the output directory exists
os.makedirs(output_directory, exist_ok=True)

# Strings to be cut off during re-search
cutoff_strings = [
    " се", " [в]", " си", " [с]", " за", " в", " (се)", " (се) [с]", " (се) да"
]

# Summary data for logging
summary_data = Counter()

# Setup logging
log_file = os.path.join(base_directory, f"debug_{datetime.now().strftime('%Y%m%dT%H%M%S')}.log")
logging.basicConfig(
    filename=log_file,
    level=logging.DEBUG,
    format="%(asctime)s - %(levelname)s - %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S"
)
logging.info("Script started")
logging.info(f"Mode: {mode}")


def extract_roms(data):
    """
    Extract ROMS from JSON data.
    """
    if not isinstance(data, dict):
        logging.error(f"extract_roms called with non-dict data type: {type(data)}; data={str(data)[:200]}")
        return []
    logging.debug(f"Entered extract_roms with data keys: {list(data.keys())}")
    hits = data.get("hits", [])
    for hit in hits:
        for rom in hit.get("roms", []):
            logging.debug(f"Yielding rom from hit: {rom}")
            yield rom
    logging.debug("Exiting extract_roms")


def extract_wordclass(rom):
    """
    Extract wordclass from headword_full.
    """
    headword_full = rom.get("headword_full", "")
    logging.debug(f"Entered extract_wordclass with headword_full: {headword_full}")
    match = re.search(r'<span class="wordclass">([^<]+)</span>', headword_full)
    if match:
        logging.debug(f"Extracted wordclass: {match.group(1)}")
    else:
        logging.debug("No wordclass found in headword_full.")
    logging.debug("Exiting extract_wordclass")
    return match.group(1) if match else None


def match_partial(bulgarian_1, data):
    """
    Match partial fields like indirect references, full_collocation, or reflection.
    """
    logging.debug(f"Entered match_partial for bulgarian_1: {bulgarian_1}")
    for rom in extract_roms(data):
        header = rom.get("header", "")
        source = rom.get("source", "")
        logging.debug(f"Checking rom header: {header}, source: {source}")

        partial_matches = [
            (r'<span class="indirect_reference_OTHER">([^<]+)</span>', "3b"),
            (r'<span class="indirect_reference_RQ">([^<]+)</span>', "3c"),
            (r'<span class="full_collocation">([^<]+)</span>', "3d"),
            (r'<span class="reflection">([^<]+)</span>', "3e"),
        ]

        for pattern, level in partial_matches:
            match = re.search(pattern, header) or re.search(pattern, source)
            if match and bulgarian_1 == match.group(1):
                logging.debug(f"Partial match found: Level {level}, Match: {match.group(1)}")
                logging.debug("Exiting match_partial with result")
                return level, match.group(1)

    logging.debug("No partial match found in match_partial.")
    logging.debug("Exiting match_partial with no match")
    return "No Match", None


def apply_cutoff_logic(bulgarian_1):
    """
    Apply cutoff logic to revise Bulgarian 1 and attempt a match.
    """
    logging.debug(f"Entered apply_cutoff_logic for bulgarian_1: {bulgarian_1}")
    for cutoff in cutoff_strings:
        if bulgarian_1.endswith(cutoff):
            revised_query = bulgarian_1[: -len(cutoff)].strip()
            logging.debug(f"Cutoff applied: {cutoff}, Revised Query: {revised_query}")
            logging.debug("Exiting apply_cutoff_logic with cutoff applied")
            return cutoff, revised_query
    logging.debug("No cutoff applied in apply_cutoff_logic.")
    logging.debug("Exiting apply_cutoff_logic with no cutoff")
    return None, None


def extract_hints(data):
    """
    Extract example hints or other supporting info from JSON data.
    """
    logging.debug(f"Entered extract_hints with data type: {type(data)}")
    hints = []
    try:
        for rom in extract_roms(data):
            examples = rom.get("examples", [])
            for example in examples:
                hints.append(example)
        logging.debug(f"Extracted hints: {hints}")
    except Exception as e:
        logging.error(f"Exception in extract_hints: {e}", exc_info=True)
    logging.debug("Exiting extract_hints")
    return hints


def fetch_and_concatenate():
    """
    Fetches data from the PONS API for each query term and concatenates all results into a single JSON file.
    """
    logging.info("Starting fetch_and_concatenate process.")
    concatenated_data = []

    try:
        with open(input_file_path, 'r', encoding='utf-8') as file:
            for idx, line in enumerate(file):
                query_term = line.strip()
                if query_term:
                    logging.debug(f"[{idx}] Fetching data for query: {query_term}")
                    url = f"https://api.pons.com/v1/dictionary?q={query_term}&l=bgen"
                    headers = {
                        "X-Secret": "XXX"
                    }
                    try:
                        response = requests.get(url, headers=headers)
                        if response.status_code == 200:
                            logging.debug(f"Successful API response for query: {query_term}")
                            concatenated_data.append({
                                "query": query_term,
                                "data": response.json()
                            })
                        else:
                            logging.warning(f"Failed API response for query: {query_term}, Status Code: {response.status_code}")
                            concatenated_data.append({
                                "query": query_term,
                                "data": {
                                    "error": f"Received status code {response.status_code}",
                                    "response_text": response.text
                                }
                            })
                    except Exception as e:
                        logging.error(f"Exception during request for {query_term}: {e}", exc_info=True)
                        concatenated_data.append({
                            "query": query_term,
                            "data": {
                                "error": f"Exception: {e}"
                            }
                        })
        # Save the concatenated data to a single JSON file
        with open(concatenated_file_path, 'w', encoding='utf-8') as output_file:
            json.dump(concatenated_data, output_file, ensure_ascii=False, indent=4)
            logging.info(f"All data fetched and concatenated into {concatenated_file_path}")
    except Exception as e:
        logging.error(f"Exception in fetch_and_concatenate: {e}", exc_info=True)


def write_results_to_xlsm(results, xlsm_path, sheet_name="Results"):
    """
    Write the results to a 'Results' worksheet in the existing XLSM file, preserving macros.
    If the sheet exists, it will be replaced.
    """
    logging.info(f"Writing {len(results)} results to '{sheet_name}' in {xlsm_path}")
    df = pd.DataFrame(results)
    wb = load_workbook(xlsm_path, keep_vba=True)
    # Remove old Results sheet if it exists
    if sheet_name in wb.sheetnames:
        std = wb[sheet_name]
        wb.remove(std)
    # Add DataFrame as new worksheet
    ws = wb.create_sheet(title=sheet_name)
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    wb.save(xlsm_path)
    logging.info(f"Wrote {len(results)} results to {xlsm_path} in sheet '{sheet_name}'")


def process_and_reconcile():
    """
    Processes entries from concatenated.json, reconciles them with Flashcards.xlsm,
    and saves the combined results into the 'Results' worksheet of Flashcards.xlsm.
    """
    logging.info("Starting process_and_reconcile function.")

    if not os.path.exists(concatenated_file_path):
        logging.error(f"Concatenated JSON file not found at {concatenated_file_path}.")
        return

    if not os.path.exists(flashcards_xlsm_path):
        logging.error(f"Flashcards.xlsm file not found at {flashcards_xlsm_path}.")
        return

    try:
        logging.info("Loading concatenated.json file.")
        with open(concatenated_file_path, 'r', encoding='utf-8') as file:
            concatenated_data = json.load(file)
        logging.debug(f"Loaded {len(concatenated_data)} entries from concatenated.json.")

        # Use pandas to read the Anki sheet from XLSM
        anki_data = []
        try:
            wb = load_workbook(flashcards_xlsm_path, read_only=True, keep_vba=True)
            if "Anki" not in wb.sheetnames:
                logging.error("Sheet 'Anki' not found in Flashcards.xlsm.")
                return
            ws = wb["Anki"]
            rows = list(ws.iter_rows(values_only=True))
            headers = rows[0]
            for i, row in enumerate(rows[1:]):
                row_data = dict(zip(headers, row))
                # Only keep relevant fields
                anki_data.append({
                    "Bulgarian 1": row_data.get("Bulgarian 1"),
                    "Bulgarian 2": row_data.get("Bulgarian 2"),
                    "Part of Speech": row_data.get("Part of Speech"),
                    "Note ID": row_data.get("Note ID")
                })
                if i % 1000 == 0:
                    logging.debug(f"Loaded row {i}: {anki_data[-1]}")
            logging.info(f"Collected {len(anki_data)} rows from Anki worksheet.")
        except Exception as e:
            logging.error(f"Error loading Anki sheet from Flashcards.xlsm: {e}", exc_info=True)
            return

        results = []

        logging.info("Starting matching logic.")
        for idx, anki_row in enumerate(anki_data):
            bulgarian_1 = anki_row["Bulgarian 1"]
            part_of_speech = anki_row["Part of Speech"]
            note_id = anki_row["Note ID"]
            logging.debug(f"Processing row {idx}: Note ID={note_id}, Bulgarian 1={bulgarian_1}, Part of Speech={part_of_speech}")

            match_level = "No Match"
            matched_value = None
            cutoff_applied = None
            hint_1 = None
            hint_2 = None
            pons_status_1 = "Unmatched"
            pons_status_2 = ""

            found_match = False

            # Try all entries in concatenated_data
            for json_entry in concatenated_data:
                query = json_entry["query"]
                data = json_entry.get("data", {})

                # Level 1: Exact match with matching wordclass
                if bulgarian_1 == query:
                    logging.debug(f"Level 1: Exact match for query: {query} (row {idx})")
                    for rom in extract_roms(data):
                        wordclass = extract_wordclass(rom)
                        logging.debug(f"Checking wordclass: {wordclass} (expected: {part_of_speech})")
                        if wordclass == part_of_speech:
                            match_level = "1"
                            matched_value = query
                            hints = extract_hints(data)
                            hint_1, hint_2 = hints if len(hints) > 1 else (hints[0], None) if hints else (None, None)
                            pons_status_1 = "Exact Match"
                            pons_status_2 = "Wordclass Match"
                            found_match = True
                            logging.debug(f"Level 1 match succeeded for {bulgarian_1}. Hints: {hint_1}, {hint_2}")
                            break
                    if found_match:
                        break

            # If not found, try Level 2
            if not found_match:
                for json_entry in concatenated_data:
                    query = json_entry["query"]
                    data = json_entry.get("data", {})
                    if bulgarian_1 == query:
                        logging.debug(f"Level 2: Exact match for query: {query} (row {idx})")
                        match_level = "2"
                        matched_value = query
                        hints = extract_hints(data)
                        hint_1, hint_2 = hints if len(hints) > 1 else (hints[0], None) if hints else (None, None)
                        pons_status_1 = "Exact Match"
                        pons_status_2 = "No Wordclass Match"
                        found_match = True
                        logging.debug(f"Level 2 match succeeded for {bulgarian_1}. Hints: {hint_1}, {hint_2}")
                        break

            # If not found, try partial match
            if not found_match:
                for json_entry in concatenated_data:
                    query = json_entry["query"]
                    data = json_entry.get("data", {})
                    level, match_val = match_partial(bulgarian_1, data)
                    if level != "No Match":
                        match_level = level
                        matched_value = match_val
                        hints = extract_hints(data)
                        hint_1, hint_2 = hints if len(hints) > 1 else (hints[0], None) if hints else (None, None)
                        pons_status_1 = "Partial Match"
                        pons_status_2 = f"Level {level}"
                        found_match = True
                        logging.debug(f"Partial match succeeded at level {level} for {bulgarian_1}. Hints: {hint_1}, {hint_2}")
                        break

            # If still not found, try cutoff logic
            if not found_match:
                cutoff, revised_query = apply_cutoff_logic(bulgarian_1)
                if cutoff and revised_query:
                    for json_entry in concatenated_data:
                        query = json_entry["query"]
                        data = json_entry.get("data", {})
                        # Try Level 1 again with revised query
                        if revised_query == query:
                            logging.debug(f"Level 4: Cutoff match for revised query: {revised_query} (cutoff: {cutoff}, row {idx})")
                            for rom in extract_roms(data):
                                wordclass = extract_wordclass(rom)
                                if wordclass == part_of_speech:
                                    match_level = "4"
                                    matched_value = revised_query
                                    cutoff_applied = cutoff
                                    hints = extract_hints(data)
                                    hint_1, hint_2 = hints if len(hints) > 1 else (hints[0], None) if hints else (None, None)
                                    pons_status_1 = "Cutoff Match"
                                    pons_status_2 = f"Wordclass Match (cutoff: {cutoff})"
                                    found_match = True
                                    logging.debug(f"Cutoff Level 4 match succeeded for {bulgarian_1} -> {revised_query}. Hints: {hint_1}, {hint_2}")
                                    break
                            if found_match:
                                break
                    # Try Level 2 with cutoff
                    if not found_match:
                        for json_entry in concatenated_data:
                            query = json_entry["query"]
                            data = json_entry.get("data", {})
                            if revised_query == query:
                                match_level = "4"
                                matched_value = revised_query
                                cutoff_applied = cutoff
                                hints = extract_hints(data)
                                hint_1, hint_2 = hints if len(hints) > 1 else (hints[0], None) if hints else (None, None)
                                pons_status_1 = "Cutoff Match"
                                pons_status_2 = f"No Wordclass Match (cutoff: {cutoff})"
                                found_match = True
                                logging.debug(f"Cutoff Level 4 match (no wordclass) succeeded for {bulgarian_1} -> {revised_query}. Hints: {hint_1}, {hint_2}")
                                break

            if not found_match:
                logging.debug(f"No match found for Note ID={note_id}, Bulgarian 1={bulgarian_1}")

            # Append result details for this row
            results.append({
                "Note ID": note_id,
                "Bulgarian 1": bulgarian_1,
                "Part of Speech": part_of_speech,
                "Match Level": match_level,
                "Matched Value": matched_value,
                "Cutoff Applied": cutoff_applied,
                "Hint 1": hint_1,
                "Hint 2": hint_2,
                "PONS Status 1": pons_status_1,
                "PONS Status 2": pons_status_2
            })

            if idx % 100 == 0 or not found_match:
                logging.debug(f"Result appended for row {idx}: Note ID={note_id}, Match Level={match_level}")

        # Save results to XLSM Results worksheet
        try:
            write_results_to_xlsm(results, flashcards_xlsm_path, sheet_name="Results")
        except Exception as e:
            logging.error(f"Exception saving results to Excel: {e}", exc_info=True)

        # Log summary statistics
        logging.info(f"Processing complete. Total rows: {len(anki_data)}. Results saved to Results worksheet in Flashcards.xlsm.")
        match_levels = Counter(result['Match Level'] for result in results)
        for level, count in match_levels.items():
            logging.info(f"Rows matched at level {level}: {count}")

    except Exception as e:
        logging.error(f"An error occurred in process_and_reconcile: {e}", exc_info=True)


# Main workflow
if mode == "fetch":
    logging.info("Main: Running fetch_and_concatenate()")
    fetch_and_concatenate()
elif mode == "process":
    logging.info("Main: Running process_and_reconcile()")
    process_and_reconcile()
else:
    logging.error(f"Unknown mode: {mode}")
    print(f"Unknown mode: {mode}")
    print("Available modes: fetch, process")
